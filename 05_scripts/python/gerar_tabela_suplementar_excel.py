import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv('09_resultados/artigo_igro_figuras_tabelas/tabelas/tabela_suplementar_igro_51_orgaos.csv')
df['orgao'] = df['orgao'].str.replace('GOI�SFOMENTO', 'GOIASFOMENTO', regex=False)

col_labels = {
    'orgao': 'Orgao',
    'classe_operacional': 'Classe',
    'manifestacoes': 'Manifestacoes (n)',
    'pesquisas': 'Pesquisas (n)',
    'tmr_dias': 'TMR (dias)',
    'pma_pct': 'PMA (%)',
    'rp_pct': 'RP (%)',
    'ri_pct': '%RI (%)',
    'nr_nps': 'NR (NPS)',
    'igro_pct': 'IGRO (%)',
    'faixa_risco': 'Faixa de Risco',
    'amostra_insuficiente': 'Amostra Insuficiente'
}
df_display = df.rename(columns=col_labels)
df_display['Amostra Insuficiente'] = df_display['Amostra Insuficiente'].map({True: 'Sim', False: 'Nao'})
df_display.insert(0, 'Ranking', range(1, len(df_display) + 1))

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cor_header     = PatternFill('solid', fgColor='1D3557')
cor_controlado = PatternFill('solid', fgColor='D4EDDA')
cor_atencao    = PatternFill('solid', fgColor='FFF3CD')
cor_elevado    = PatternFill('solid', fgColor='FFE0B2')
cor_critico    = PatternFill('solid', fgColor='F8D7DA')
faixa_cores = {'Controlado': cor_controlado, 'Atencao': cor_atencao,
               'Elevado': cor_elevado, 'Critico': cor_critico}

wb = Workbook()

# ── ABA 1: Resultados ──────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Resultados IGRO'

ws.merge_cells('A1:M1')
ws['A1'] = 'Tabela Suplementar - Ranking IGRO: Resultados por Orgao (2024-2025)'
ws['A1'].font = Font(bold=True, size=13, color='FFFFFF', name='Garamond')
ws['A1'].fill = cor_header
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 22

ws.merge_cells('A2:M2')
ws['A2'] = 'Fonte: SGOe (2024-2025). Elaboracao propria. N = 51 orgaos do Poder Executivo do Estado de Goias.'
ws['A2'].font = Font(italic=True, size=10, name='Garamond')
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 16

headers = list(df_display.columns)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', name='Garamond', size=10)
    cell.fill = cor_header
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[3].height = 30

numeric_cols = {'TMR (dias)', 'PMA (%)', 'RP (%)', '%RI (%)', 'NR (NPS)', 'IGRO (%)'}
for row_idx, row in df_display.iterrows():
    excel_row = row_idx + 4
    faixa = df.iloc[row_idx]['faixa_risco']
    row_fill = faixa_cores.get(faixa, PatternFill('solid', fgColor='F8F9FA'))
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=excel_row, column=col_idx, value=value)
        cell.font = Font(name='Garamond', size=10)
        cell.fill = row_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if headers[col_idx - 1] in numeric_cols:
            cell.number_format = '0.0'
    ws.row_dimensions[excel_row].height = 15

col_widths = [8, 22, 7, 18, 14, 9, 8, 8, 8, 10, 10, 15, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A4'

# ── ABA 2: Dicionario de Dados ─────────────────────────────────────────────
ws2 = wb.create_sheet('Dicionario de Dados')
dic_header_fill = PatternFill('solid', fgColor='457B9D')
dic_alt = PatternFill('solid', fgColor='EBF5FB')

ws2.merge_cells('A1:F1')
ws2['A1'] = 'Dicionario de Dados - Tabela Suplementar IGRO'
ws2['A1'].font = Font(bold=True, size=13, color='FFFFFF', name='Garamond')
ws2['A1'].fill = cor_header
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 22

dic_headers = ['Variavel (original)', 'Rotulo', 'Tipo', 'Unidade / Escala', 'Descricao', 'Fonte']
for col_idx, h in enumerate(dic_headers, 1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF', name='Garamond', size=10)
    cell.fill = dic_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws2.row_dimensions[2].height = 22

dicionario = [
    ('ranking', 'Ranking', 'Inteiro', '1 a 52',
     'Posicao do orgao no ranking IGRO, do maior (1) para o menor.',
     'Calculado'),
    ('orgao', 'Orgao', 'Texto', '-',
     'Sigla do orgao publico estadual conforme cadastro do SGOe.',
     'SGOe'),
    ('classe_operacional', 'Classe', 'Inteiro', '1 a 5',
     'Classe operacional: Cl.1 (>10.000 manifest.), Cl.2 (2.001-10.000), Cl.3 (501-2.000), Cl.4 (51-500), Cl.5 (<=50).',
     'Elaboracao propria'),
    ('manifestacoes', 'Manifestacoes (n)', 'Inteiro', 'Unidade',
     'Total de manifestacoes cidadas registradas no orgao entre jan/2024 e dez/2025.',
     'SGOe'),
    ('pesquisas', 'Pesquisas (n)', 'Inteiro', 'Unidade',
     'Total de respostas validas recebidas na pesquisa de satisfacao pos-atendimento.',
     'SGOe'),
    ('tmr_dias', 'TMR (dias)', 'Decimal', 'Dias',
     'Tempo Medio de Resposta: media de dias entre o registro da manifestacao e a resposta definitiva ao cidadao.',
     'SGOe'),
    ('pma_pct', 'PMA (%)', 'Decimal', 'Percentual (0-100)',
     'Percentual de Manifestacoes em Atraso: proporcao respondidas apos o prazo legal de 30 dias.',
     'SGOe'),
    ('rp_pct', 'RP (%)', 'Decimal', 'Percentual (0-100)',
     'Resolutividade Percebida: percentual de respondentes que avaliaram a manifestacao como resolvida.',
     'SGOe / Pesquisa'),
    ('ri_pct', '%RI (%)', 'Decimal', 'Percentual (0-100)',
     'Percentual de Respostas Insatisfatorias: proporcao de manifestacoes reabertas apos encerramento.',
     'SGOe'),
    ('nr_nps', 'NR (NPS)', 'Decimal', 'Escala -100 a +100',
     'Nota de Recomendacao (Net Promoter Score). Positivo = mais promotores que detratores.',
     'SGOe / Pesquisa'),
    ('igro_pct', 'IGRO (%)', 'Decimal', 'Percentual (0-100)',
     'Indice de Gestao de Riscos de Ouvidoria: media geometrica ponderada Sub-IGRO_T (40%) e Sub-IGRO_Q (60%). '
     'Proximo de 100 = baixo risco; proximo de 0 = risco critico.',
     'Elaboracao propria'),
    ('faixa_risco', 'Faixa de Risco', 'Texto', 'Categorico',
     'Classificacao semafórica: Controlado (>=80%), Atencao (70-79%), Elevado (50-69%), Critico (<50%).',
     'Elaboracao propria'),
    ('amostra_insuficiente', 'Amostra Insuficiente', 'Booleano', 'Sim / Nao',
     'Indica se o numero de respondentes foi inferior a 30 (n<30), limiar abaixo do qual RP e NR tem confiabilidade reduzida.',
     'Elaboracao propria'),
]

for row_idx, row_data in enumerate(dicionario, 3):
    fill = dic_alt if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='Garamond', size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True,
                                   horizontal='center' if col_idx <= 4 else 'left')
    ws2.row_dimensions[row_idx].height = 50

for i, w in enumerate([22, 22, 12, 22, 60, 20], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A3'

out = '09_resultados/artigo_igro_figuras_tabelas/tabelas/tabela_suplementar_igro_51_orgaos.xlsx'
wb.save(out)
print(f'Salvo: {out}')
